"""Export service for generating Excel reports."""

import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)


def export_orders_to_excel(analytics_data):
    """
    Generate an Excel workbook from processed analytics data.

    Args:
        analytics_data (dict): Processed analytics from OrderService.process_orders().

    Returns:
        io.BytesIO: In-memory Excel file.
    """
    wb = Workbook()

    # --- Summary Sheet ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    header_font = Font(bold=True, size=14, color='4F46E5')
    label_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='1E1B4B', end_color='1E1B4B', fill_type='solid')
    header_text = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws_summary['A1'] = 'OrderPulse Analytics Report'
    ws_summary['A1'].font = Font(bold=True, size=18, color='6366F1')

    summary = analytics_data.get('summary', {})
    row = 3
    metrics = [
        ('Total Orders', summary.get('total_orders', 0)),
        ('AWB Assigned', summary.get('awb_count', 0)),
        ('Avg Dispatch Time (hrs)', summary.get('avg_dispatch_time', 0)),
        ('Within 24h Rate (%)', summary.get('within_24h_rate', 0))
    ]
    for label, value in metrics:
        ws_summary.cell(row=row, column=1, value=label).font = label_font
        ws_summary.cell(row=row, column=2, value=value)
        row += 1

    # Store distribution
    row += 1
    ws_summary.cell(row=row, column=1, value='Store Distribution').font = header_font
    row += 1
    dist = analytics_data.get('store_distribution', {})
    for store, count in dist.items():
        ws_summary.cell(row=row, column=1, value=store).font = label_font
        ws_summary.cell(row=row, column=2, value=count)
        row += 1

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15

    # --- Orders Sheet ---
    ws_orders = wb.create_sheet('Orders')
    headers = ['Order Code', 'Store Type', 'AWB Number', 'Dispatch Time (hrs)', 'Status']

    for col, header in enumerate(headers, 1):
        cell = ws_orders.cell(row=1, column=col, value=header)
        cell.font = header_text
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    table_data = analytics_data.get('table_data', [])
    for row_idx, order in enumerate(table_data, 2):
        values = [
            order.get('code', ''),
            order.get('store_type', ''),
            order.get('awb', ''),
            order.get('dispatch_time', ''),
            order.get('status', '')
        ]
        for col, value in enumerate(values, 1):
            cell = ws_orders.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    for col in range(1, 6):
        ws_orders.column_dimensions[chr(64 + col)].width = 20

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(f'Generated Excel export with {len(table_data)} orders.')
    return output
